import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Border, Side, PatternFill

#First of all make the trial with two column HOA & Amount

file_path = r'E:\OneDrive\Desktop\trial2.xlsx'
df = pd.read_excel(file_path)

df = df.dropna()
df =df.astype({"HOA" : 'int64'})

expenses_dict = {
    'advertisement_admn': [3796, 3702],
    'advertisement_storage': [3798],
    'commemoration': [3675],
    'audit_fees': [3461, 3462, 3463, 3464, 3465, 3466, 3467],
    'bank_commission': [3722],
    'books_periodicals': [3703, 3704],
    'computer_applications': [3764],
    'expenses_incurred_on_covid': [3674],
    'computer_stationery': [3750],
    'daily_wages': [3761, 3720],
    'data_processing_communications': [3751],
    'electricity': [3382],
    'entertainment': [3705],
    'exp_on_recruitment_through_agencies': [3760],
    'exp_on_gift_for_parliamentary_consult_committee': [3770, 3718],
    'exp_on_hindi_promotion': [3709],
    'exp_on_tea_coffee_etc_within_office_premises': [3748],
    'exp_on_tea_coffee_outside_office_premises': [3749],
    'exp_other_than_gift_for_parliam_consult_committee': [3719, 3772],
    'expenditure_incurred_by_mo': [3651],
    'expenses_on_cisf_other_than_depot': [3712, 3721],
    'expenses_on_training': [3713],
    'inspection_charges_on_gis': [3737],
    'inspection_charges_on_rpfc': [3765, 3766, 3767, 3793],
    'insurance': [3353, 3354, 3355, 3356, 3357, 3358],
    'legal_fees': [3706, 3769],
    'maintenance_of_computer': [3535],
    'maintenance_of_vehicles': [3551],
    'maintenance_of_holiday_homesguest_houses': [3716],
    'medical_fixed': [6096],
    'medical_others': [6091, 6092, 6093, 6094, 6095],
    'ota_for_districtdepot': [6077],
    'ota_for_hqrozo': [6077],
    'other_sundry_expenses': [3791],
    'photostat_charges': [3756],
    'postage': [3728],
    'pr_publicity': [3671, 3672, 3673],
    'plantation_of_trees': [3768],
    'printing_stationery': [3739],
    'refund_of_unclaimed_amount': [3746],
    'reimb_of_travelling_exp_etc_to_retired_employees': [3726],
    'reimb_of_local_conveyance': [3731],
    'internal_audit_outsource': [],
    'remun_for_hiring_services_for_contingent_jobs': [3733],
    'rent': [3201, 3202, 3203, 3204, 3205, 3253, 3254, 3255, 3291],
    'research_development': [3711],
    'repair_maintenance': [3531, 3533, 3534],
    'stipend_paid_to_trainees_of_icai': [3792],
    'stipend_to_mgmt_trainees': [3752],
    'sundry_articles_for_office_use': [3714],
    'swactha_action_plan': [3795],
    'taxes': [3301, 3302, 3303, 3305, 3306, 3307, 3308, 3311, 3312, 3331, 3332, 3771, 3773, 3774, 3775, 6136],
    'taxi_hiring_charges': [3754],
    'telephone_incl_fax_mobile': [3730],
    'transport_charges_for__sundry_articles': [3727],
    'travelling_exp_on_mgmt_trainees': [3753],
    'travelling_expenses': [3411, 3412, 3413, 3414, 3415, 3416, 3417, 3418, 3419, 3431, 3432, 3441, 3442, 3445, 3451, 3452, 3453],
    'water_charges': [3755],
    'welfare': [3101, 3102, 3103, 3104, 3105, 3106, 3107, 3108, 3109, 3110, 3111, 3112, 3113, 3114, 3115, 3768],
    'furniture_fittings':[2161],
    'other_equipment':[2149]
}
budget_dict = {
    'advertisement_admn': 0,
    'advertisement_storage': 0,
    'commemoration': 0,
    'audit_fees': 0,
    'bank_commission': 0,
    'books_periodicals': 5000,
    'computer_applications': 0,
    'expenses_incurred_on_covid': 0,
    'computer_stationery': 20000,
    'daily_wages': 200000,
    'data_processing_communications': 0,
    'electricity': 120000,
    'entertainment': 0,
    'exp_on_recruitment_through_agencies': 0,
    'exp_on_gift_for_parliamentary_consult_committee': 0,
    'exp_on_hindi_promotion': 14000,
    'exp_on_tea_coffee_etc_within_office_premises': 10000,
    'exp_on_tea_coffee_outside_office_premises': 0,
    'exp_other_than_gift_for_parliam_consult_committee': 0,
    'expenditure_incurred_by_mo': 0,
    'expenses_on_cisf_other_than_depot': 0,
    'expenses_on_training': 0,
    'inspection_charges_on_gis': 0,
    'inspection_charges_on_rpfc': 0,
    'insurance': 0,
    'legal_fees': 50000,
    'maintenance_of_computer': 35000,
    'maintenance_of_vehicles': 0,
    'maintenance_of_holiday_homesguest_houses': 0,
    'medical_fixed': 0,
    'medical_others': 500000,
    'ota_for_districtdepot': 400000,
    'ota_for_hqrozo': 0,
    'other_sundry_expenses': 2000,
    'photostat_charges': 20000,
    'postage': 10000,
    'pr_publicity': 0,
    'plantation_of_trees': 0,
    'printing_stationery': 50000,
    'refund_of_unclaimed_amount': 0,
    'reimb_of_travelling_exp_etc_to_retired_employees': 0,
    'reimb_of_local_conveyance': 5000,
    'internal_audit_outsource': 0,
    'remun_for_hiring_services_for_contingent_jobs': 10000,
    'rent': 500000,
    'research_development': 0,
    'repair_maintenance': 10000,
    'stipend_paid_to_trainees_of_icai': 0,
    'stipend_to_mgmt_trainees': 0,
    'sundry_articles_for_office_use': 10000,
    'swactha_action_plan': 6000,
    'taxes': 100000,
    'taxi_hiring_charges': 200000,
    'telephone_incl_fax_mobile': 50000,
    'transport_charges_for__sundry_articles': 0,
    'travelling_exp_on_mgmt_trainees': 0,
    'travelling_expenses': 600000,
    'water_charges': 5000,
    'welfare': 150000,
    'furniture_fittings':50000,
    'other_equipment':100000
}


expense_sums = {}
hoa_lists = {}

for category, hoa_list in expenses_dict.items():
    if hoa_list:
        filtered_data = df[df['HOA'].isin(hoa_list)]
        total_amount = filtered_data['Amount'].sum()
        expense_sums[category] = total_amount
        hoa_lists[category] = hoa_list  # Store the HOA list for the category
    else:
        expense_sums[category] = 0
        hoa_lists[category] = []  # Store an empty list if there are no HOAs

# Create DataFrame with sums, HOA lists, budgets, and differences
sums_df = pd.DataFrame({
    'Expense Category': expense_sums.keys(),
    'Head of Account': [hoa_lists[category] for category in expense_sums.keys()],
    'Total Amount': expense_sums.values(),
    'Budget': [budget_dict.get(category, 0) for category in expense_sums.keys()],
})

# Ensure numeric types for comparison
sums_df['Total Amount'] = pd.to_numeric(sums_df['Total Amount'], errors='coerce')
sums_df['Budget'] = pd.to_numeric(sums_df['Budget'], errors='coerce')

# Calculate the difference and add it as a new column
#sums_df['Difference'] = sums_df['Total Amount'] - sums_df['Budget']
sums_df['Difference'] = sums_df['Budget'] - sums_df['Total Amount']


output_file_path = r'E:\OneDrive\Desktop\expense_sums.xlsx'
sums_df.to_excel(output_file_path, index=False)

# Load the workbook and add borders
workbook = load_workbook(output_file_path)
sheet = workbook.active

# Define border style
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# # Apply border to each cell in the DataFrame
# for row in sheet.iter_rows(min_row=1, max_row=sums_df.shape[0] + 1, min_col=1, max_col=sums_df.shape[1]):
#     for cell in row:
#         cell.border = border
# Define fill style for red cells
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Apply border and color to each cell in the DataFrame
for row in sheet.iter_rows(min_row=1, max_row=sums_df.shape[0] + 1, min_col=1, max_col=sums_df.shape[1]):
    for cell in row:
        cell.border = border
        # Highlight cells in red if Total Amount exceeds Budget
        if cell.column == 3:  # Total Amount column (index 3)
            total_amount = pd.to_numeric(cell.value, errors='coerce')  # Ensure it's numeric
            budget = pd.to_numeric(sums_df.iloc[cell.row - 2]['Budget'], errors='coerce')  # Ensure it's numeric
            if pd.notna(total_amount) and pd.notna(budget) and total_amount > budget:  # Compare with corresponding budget
                cell.fill = red_fill
# Save the workbook
workbook.save(output_file_path)
print(f"Expense sums exported to {output_file_path} with borders.")


#C:\Users\Abcd\PycharmProjects\overhead>python overhead2.py
#Expense sums exported to E:\OneDrive\Desktop\expense_sums.xlsx with borders.

#C:\Users\Abcd\PycharmProjects\overhead>

# df = pd.read_excel(r'C:\Users\kusha\Desktop\newdf3.xlsx')
#
# df = df.to_excel(r'C:\Users\kusha\Desktop\newdf3.xlsx')
#
# delete the first/any column
# del df[df.columns[0]]
# df = df.drop(df.columns[[0,1]],axis=1)
#
# delete n rows after selecting all rows after
# df = df.iloc[10:,:]
#
# set column name of row index 0
# df.columns = df.iloc[0]
#
#
# df = df[1:]
#
# df.info()
#
#
# df = df.dropna()
#
# df.drop(df.index[(df["Product"]=="Product")],axis=0,inplace=True)
#
# df = df.astype({"Account":'int64'})
#
# print(df.sum())
#
# print(df[df['Account'].isin(range(2112,2171))])
#
# sum_1=df.loc[df['Account'].isin(range(2112,2172)),'Closing Balance'].sum()
#
# sum_2=df.loc[df['Account'].isin(range(2212,2272)),'Closing Balance'].sum()
#
# sum_3=df.loc[df['Account'].isin(range(2312,2372)),'Closing Balance'].sum()
#
# sum_4=df.loc[df['Account']==2361,'Closing Balance'].sum()
#
# df = df.to_excel(r'C:\Users\kusha\Desktop\newdf3.xlsx')
# df = df.dropna()
# df = df.astype({'Account': 'int64', 'Product': 'int64', 'Scheme':'int64','Opening Balance':'float64','Debit':'float64','Credit':'float64','Closing Balance':'float64'})